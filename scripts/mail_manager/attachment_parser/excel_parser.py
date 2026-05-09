"""
Excel parser implementation using openpyxl.

Extracts text and metadata from Excel spreadsheets (.xlsx, .xls).
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class ExcelParser:
    """Parser for Excel documents using openpyxl."""

    def can_parse(self, file_path: Path) -> bool:
        """
        Check if this parser can handle Excel files.

        Args:
            file_path: Path to the file to check.

        Returns:
            True if file has .xlsx or .xls extension.
        """
        return file_path.suffix.lower() in (".xlsx", ".xls")

    def extract_text(self, file_path: Path) -> str:
        """
        Extract text content from Excel file.

        Iterates through all sheets and all rows, joining cell values
        with " | " separator.

        Args:
            file_path: Path to the Excel file.

        Returns:
            Extracted text from all sheets, joined by newlines.
        """
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            logger.error(f"Failed to open Excel {file_path}: {e}")
            return ""
        try:
            text_parts = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell else "" for cell in row)
                    if row_text.strip():
                        text_parts.append(row_text)
            return "\n".join(text_parts)
        finally:
            wb.close()

    def extract_metadata(self, file_path: Path) -> dict[str, Any]:
        """
        Extract metadata from Excel file.

        Args:
            file_path: Path to the Excel file.

        Returns:
            Dictionary with sheet_names and type.
        """
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            logger.error(f"Failed to open Excel {file_path}: {e}")
            return {"type": "excel", "sheets": []}
        try:
            return {
                "sheet_names": list(wb.sheetnames),
                "type": "excel",
            }
        finally:
            wb.close()
